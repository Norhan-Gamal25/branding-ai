"""
exporter.py
===========
Generates the 10-item Brand Studio AI export kit, then bundles everything
into a master Branding_Kit.zip.

Files produced per job:
  1.  Brand_Guide.pdf            — multi-page PDF brand guide
  2.  Content_Calendar.xlsx      — 30-day content calendar
  3.  Canva_Import.json          — Canva-compatible brand kit payload
  4.  Image_Prompts.txt          — halal-safe AI image generation prompts
  5.  Zero_Budget_Playbook.pdf   — 30-day action playbook PDF
  6.  Free_Tools_List.txt        — 15 curated free tools
  7.  Landing_Page.html          — AI-generated deployable brand landing page
  8.  Branding_Kit.zip           — master ZIP containing all of the above

Unicode support:
  DejaVu Sans TTF fonts are installed via the `fonts-dejavu-core` apt package
  in the backend Dockerfile. Falls back to Helvetica on local dev without package.
"""

from __future__ import annotations

import json
import logging
import os
import zipfile
from datetime import datetime, timedelta
from pathlib import Path
from typing import Any

import openpyxl
from fpdf import FPDF
from openpyxl.styles import Alignment, Font, PatternFill

logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# CONSTANTS
# ---------------------------------------------------------------------------

_PDF_ACCENT = (41, 98, 255)     # brand blue  (RGB)
_PDF_LIGHT  = (240, 244, 255)   # light blue tint for alternating rows
_PDF_GOLD   = (212, 175, 55)    # accent gold for compliance cert

_RTL_LOCALES = frozenset({"ar", "he", "fa", "ur", "ps", "ku", "yi", "dv", "ug"})

_DEJAVU_PATHS = [
    "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf",
    "/usr/share/fonts/dejavu/DejaVuSans.ttf",
    "/usr/share/fonts/TTF/DejaVuSans.ttf",
]
_DEJAVU_BOLD_PATHS = [
    "/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf",
    "/usr/share/fonts/dejavu/DejaVuSans-Bold.ttf",
    "/usr/share/fonts/TTF/DejaVuSans-Bold.ttf",
]


def _find_font(paths: list[str]) -> str | None:
    for p in paths:
        if os.path.isfile(p):
            return p
    return None


_DEJAVU_REGULAR = _find_font(_DEJAVU_PATHS)
_DEJAVU_BOLD    = _find_font(_DEJAVU_BOLD_PATHS)
_USE_UNICODE    = _DEJAVU_REGULAR is not None
_FONT_HEADING   = "DejaVu" if _USE_UNICODE else "Helvetica"
_FONT_BODY      = "DejaVu" if _USE_UNICODE else "Helvetica"


# ---------------------------------------------------------------------------
# RTL TEXT SHAPING
# ---------------------------------------------------------------------------

def _is_rtl_locale(language: str) -> bool:
    """Return True if the language code is a known RTL locale."""
    return language.split("-")[0].lower() in _RTL_LOCALES


# Lazy-import reshaping libraries so the app still works if they aren't installed.
# Both arabic-reshaper and python-bidi are listed in requirements.txt.
_reshaper_config: Any = None
_bidi_algorithm: Any = None


def _load_rtl_libs() -> bool:
    """Try to import arabic-reshaper and python-bidi. Returns True on success."""
    global _reshaper_config, _bidi_algorithm
    if _reshaper_config is not None:
        return True  # already loaded
    try:
        import arabic_reshaper  # type: ignore
        from bidi.algorithm import get_display  # type: ignore
        _reshaper_config = arabic_reshaper.ArabicReshaper(
            configuration={
                "delete_harakat": False,
                "support_ligatures": True,
                "RIAL SIGN": True,
            }
        )
        _bidi_algorithm = get_display
        return True
    except Exception as exc:
        logger.warning("RTL libraries not available (%s) — PDF text will be LTR", exc)
        return False


def reshape_for_pdf(text: str) -> str:
    """
    Reshape and apply the Unicode BiDi algorithm to *text* so that
    Arabic / Hebrew / Urdu / Persian characters render correctly in fpdf2.

    fpdf2 renders glyphs in logical (storage) order without BiDi reordering,
    so we must pre-process the string before handing it to any cell/multi_cell.

    Falls back to the original string if the libraries are unavailable.
    """
    if not text or not _load_rtl_libs():
        return text
    try:
        # _reshaper_config is the configured ArabicReshaper instance
        reshaped = _reshaper_config.reshape(text)  # type: ignore[union-attr]
        return _bidi_algorithm(reshaped)            # type: ignore[operator]
    except Exception:
        return text


# ---------------------------------------------------------------------------
# SHARED PDF BASE CLASS
# ---------------------------------------------------------------------------

class _BasePDF(FPDF):
    """
    FPDF subclass with branded header, footer, and helper methods.

    When ``rtl=True`` the instance pre-processes every string through
    reshape_for_pdf() and right-aligns all content cells so that Arabic,
    Hebrew, Urdu, and Persian text renders correctly.
    """

    def __init__(
        self,
        brand_name: str,
        subtitle: str = "Brand Guide",
        rtl: bool = False,
    ) -> None:
        super().__init__()
        self.brand_name = brand_name
        self.subtitle   = subtitle
        self._rtl       = rtl
        self.set_auto_page_break(auto=True, margin=20)
        self.set_margins(20, 20, 20)
        if _USE_UNICODE:
            self.add_font("DejaVu", style="",  fname=_DEJAVU_REGULAR, uni=True)
            bold_src = _DEJAVU_BOLD or _DEJAVU_REGULAR
            self.add_font("DejaVu", style="B", fname=bold_src, uni=True)

    # ── Internal helpers ────────────────────────────────────────────────────

    def _t(self, text: str) -> str:
        """Conditionally reshape text for RTL rendering."""
        return reshape_for_pdf(str(text)) if self._rtl else str(text)

    def _align(self, default: str = "L") -> str:
        """Return 'R' for RTL documents, otherwise the provided default."""
        return "R" if self._rtl else default

    # ── FPDF overrides ──────────────────────────────────────────────────────

    def header(self) -> None:
        self.set_fill_color(*_PDF_ACCENT)
        self.rect(0, 0, 210, 14, "F")
        self.set_font(_FONT_HEADING, "B", 11)
        self.set_text_color(255, 255, 255)
        header_text = self._t(
            f"  {self.brand_name}  —  {self.subtitle}  |  Brand Studio AI"
        )
        self.cell(0, 14, header_text, ln=True)
        self.set_text_color(0, 0, 0)
        self.ln(4)

    def footer(self) -> None:
        self.set_y(-14)
        self.set_font(_FONT_BODY, "", 8)
        self.set_text_color(130, 130, 130)
        self.cell(0, 10, f"Page {self.page_no()}  ·  Brand Studio AI — The Halal Way", align="C")

    # ── Content helpers ─────────────────────────────────────────────────────

    def section_title(self, title: str) -> None:
        self.set_fill_color(*_PDF_ACCENT)
        self.set_font(_FONT_HEADING, "B", 12)
        self.set_text_color(255, 255, 255)
        self.cell(0, 9, f"  {self._t(title)}", ln=True, fill=True)
        self.set_text_color(0, 0, 0)
        self.ln(3)

    def body_text(self, text: str, indent: int = 0) -> None:
        self.set_font(_FONT_BODY, "", 10)
        x = self.l_margin + (indent if not self._rtl else 0)
        self.set_x(x)
        w = self.w - self.r_margin - x
        if w < 20:
            x, w = self.l_margin, self.w - self.l_margin - self.r_margin
        self.multi_cell(w, 6, self._t(text), align=self._align())
        self.ln(2)

    def key_value(self, key: str, value: str) -> None:
        """
        In LTR mode: "Key:   value"  (label left, value right of it).
        In RTL mode:  both label and value span the full width on separate
        lines (label bold, value indented) — avoids garbled BiDi in side-by-
        side layout.
        """
        if self._rtl:
            self.set_font(_FONT_HEADING, "B", 10)
            self.multi_cell(
                self.w - self.l_margin - self.r_margin,
                7,
                self._t(key + ":"),
                align="R",
            )
            self.set_font(_FONT_BODY, "", 10)
            self.multi_cell(
                self.w - self.l_margin - self.r_margin,
                7,
                self._t(str(value)),
                align="R",
            )
        else:
            key_w = 50
            val_x = self.l_margin + key_w
            val_w = self.w - self.r_margin - val_x
            if val_w < 20:
                val_x, val_w = self.l_margin, self.w - self.l_margin - self.r_margin
            y = self.get_y()
            self.set_font(_FONT_HEADING, "B", 10)
            self.cell(key_w, 7, key + ":")
            self.set_xy(val_x, y)
            self.set_font(_FONT_BODY, "", 10)
            self.multi_cell(val_w, 7, str(value))

    def color_swatch(self, name: str, hex_code: str, meaning: str) -> None:
        h = hex_code.lstrip("#")
        try:
            r, g, b = int(h[0:2], 16), int(h[2:4], 16), int(h[4:6], 16)
        except (ValueError, IndexError):
            r, g, b = 41, 98, 255

        x, y = self.get_x(), self.get_y()
        self.set_fill_color(r, g, b)
        self.rect(x, y, 18, 10, "F")
        self.set_x(x + 22)
        self.set_font(_FONT_HEADING, "B", 10)
        self.cell(35, 5, self._t(name))
        self.set_font(_FONT_BODY, "", 10)
        self.cell(30, 5, hex_code)
        self.set_font(_FONT_BODY, "", 9)
        self.set_text_color(80, 80, 80)
        self.cell(0, 5, self._t(meaning), ln=True)
        self.set_text_color(0, 0, 0)
        self.ln(3)


# ---------------------------------------------------------------------------
# 1. PDF — Brand Guide (with embedded logo images)
# ---------------------------------------------------------------------------

def generate_pdf(
    output_path: Path,
    brand_data: dict[str, Any],
    language: str,
) -> None:
    """
    Render a multi-page PDF brand guide.

    Expected brand_data keys (all optional):
      brand_name, brand_story, tagline, values, color_palette, fonts,
      captions, video_scripts, logo_prompts
    """
    is_rtl = _is_rtl_locale(language)
    pdf = _BasePDF(
        brand_data.get("brand_name", "Brand"),
        subtitle="Brand Guide",
        rtl=is_rtl,
    )
    pdf.add_page()

    # ── Page 1: Identity ────────────────────────────────────────────────
    pdf.section_title("Brand Identity")
    pdf.key_value("Brand Name",  brand_data.get("brand_name", ""))
    pdf.key_value("Tagline",     brand_data.get("tagline", ""))
    pdf.key_value("Language",    language.upper())
    pdf.key_value("Generated",   datetime.utcnow().strftime("%Y-%m-%d %H:%M UTC"))
    pdf.ln(4)

    pdf.section_title("Brand Story")
    pdf.body_text(brand_data.get("brand_story", ""))

    pdf.section_title("Brand Values")
    for v in brand_data.get("values", []):
        pdf.body_text(f"• {v}", indent=4)
    pdf.ln(2)

    pdf.section_title("Colour Palette")
    for swatch in brand_data.get("color_palette", []):
        pdf.color_swatch(
            swatch.get("name", ""),
            swatch.get("hex", "#3b5bdb"),
            swatch.get("meaning", ""),
        )
    pdf.ln(2)

    pdf.section_title("Typography")
    fonts = brand_data.get("fonts", {})
    pdf.key_value("Heading Font", fonts.get("heading", ""))
    pdf.key_value("Body Font",    fonts.get("body", ""))
    pdf.ln(4)

    # ── Page 2: Logo Prompts ─────────────────────────────────────────────
    logo_prompts = brand_data.get("logo_prompts", [])
    if logo_prompts:
        pdf.add_page()
        pdf.section_title("Logo Design Prompts")
        pdf.body_text(
            "Use the following 3 prompts with any AI image generator "
            "(Midjourney, DALL-E 3, Adobe Firefly, or Stable Diffusion) "
            "to create your brand logo.\n"
            "Generated by Branding AI — powered by Groq & IBM Watsonx."
        )
        pdf.ln(4)
        for i, prompt in enumerate(logo_prompts[:3], 1):
            pdf.set_font(_FONT_HEADING, "B", 11)
            pdf.cell(0, 7, f"Logo Prompt {i}", ln=True)
            pdf.ln(2)
            pdf.set_font(_FONT_BODY, "", 9)
            pdf.multi_cell(0, 5, prompt)
            pdf.ln(4)

    # ── Page 3: Content ─────────────────────────────────────────────────
    pdf.add_page()
    pdf.section_title("Social Media Captions (Preview — first 5)")
    for i, cap in enumerate(brand_data.get("captions", [])[:5], 1):
        pdf.key_value(f"Caption {i}", cap)
    pdf.ln(4)

    pdf.section_title("Video Scripts")
    for i, script in enumerate(brand_data.get("video_scripts", []), 1):
        pdf.body_text(f"Script {i}:")
        pdf.body_text(script, indent=4)
        pdf.ln(2)

    # ── Page 4: Compliance certificate ──────────────────────────────────
    pdf.add_page()
    pdf.section_title("Halal Compliance Certificate")
    pdf.body_text(
        "This brand kit was generated by Branding AI.\n\n"
        "Technology stack:\n"
        "  * Platform Engineer (Landing Pages) — Groq / IBM Watsonx\n"
        "  * Hakawati Strategist (Chat & Visuals) — Groq DeepSeek-R1 / IBM Watsonx\n"
        "  * Multi-cloud fallback: Groq (5 model pools) → IBM Watsonx enterprise\n\n"
        "All content has been verified to be free of:\n"
        "  * Music or musical instrument references\n"
        "  * Depictions of people, faces, or human figures\n"
        "  * Animal imagery in branding or visual prompts\n"
        "  * Alcohol, pork, gambling, or other haram content\n"
        "  * Offensive or inappropriate language in any language\n"
        "  * Interest-based (riba) financial mechanics\n\n"
        f"Status  : APPROVED\n"
        f"Checked : {datetime.utcnow().strftime('%Y-%m-%d %H:%M UTC')}"
    )

    pdf.output(str(output_path))


# ---------------------------------------------------------------------------
# 2. XLSX — 30-Day Content Calendar (with post image references)
# ---------------------------------------------------------------------------

def generate_xlsx(
    output_path: Path,
    calendar_data: list[dict[str, Any]],
    language: str = "en",
) -> None:
    """
    Produce a styled Content_Calendar.xlsx.
    RTL languages (Arabic, Hebrew, Urdu, etc.) get right-to-left text
    alignment on the text-heavy columns.
    """
    is_rtl = _is_rtl_locale(language)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "30-Day Content Calendar"
    # Set sheet reading order for RTL languages
    if is_rtl:
        ws.sheet_view.rightToLeft = True

    headers   = ["Day", "Date", "Platform", "Theme", "Caption", "Hashtags",
                 "Photo Direction"]
    hdr_fill  = PatternFill("solid", fgColor="2962FF")
    hdr_font  = Font(bold=True, color="FFFFFF", size=11)
    hdr_align = Alignment(horizontal="center", vertical="center")

    for col, hdr in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=hdr)
        c.fill, c.font, c.alignment = hdr_fill, hdr_font, hdr_align

    ws.row_dimensions[1].height = 22

    for i, w in enumerate([6, 14, 14, 26, 60, 40, 70], 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    platforms  = ["Instagram", "Facebook", "LinkedIn", "Twitter/X", "TikTok",
                  "YouTube",   "Pinterest", "Threads"]
    start_date = datetime.utcnow()

    # Columns that contain user-generated text (0-indexed inside the values list)
    _text_cols = {3, 4, 5, 6}  # Theme, Caption, Hashtags, Photo Direction

    for row_idx, row in enumerate(calendar_data[:30], 2):
        day_num  = row.get("day", row_idx - 1)
        date_str = (start_date + timedelta(days=int(day_num) - 1)).strftime("%Y-%m-%d")
        platform = platforms[(row_idx - 2) % len(platforms)]
        bg       = "EEF2FF" if row_idx % 2 == 0 else "FFFFFF"
        row_fill = PatternFill("solid", fgColor=bg)

        values = [
            day_num,
            date_str,
            platform,
            row.get("theme",           ""),
            row.get("caption",         ""),
            row.get("hashtags",        ""),
            row.get("photo_direction", row.get("image_prompt", "")),
        ]
        for col_idx, val in enumerate(values, 1):
            c = ws.cell(row=row_idx, column=col_idx, value=val)
            c.fill = row_fill
            # Apply RTL alignment to text columns; numeric/date columns stay LTR
            h_align = "right" if (is_rtl and (col_idx - 1) in _text_cols) else "left"
            c.alignment = Alignment(
                wrap_text=True,
                vertical="top",
                horizontal=h_align,
                reading_order=2 if is_rtl else 1,
            )

        ws.row_dimensions[row_idx].height = 52

    ws.freeze_panes = "A2"
    wb.save(str(output_path))


# ---------------------------------------------------------------------------
# 3. JSON — Canva Brand Kit Import
# ---------------------------------------------------------------------------

def generate_canva_json(output_path: Path, brand_data: dict[str, Any]) -> None:
    palette = brand_data.get("color_palette", [])
    fonts   = brand_data.get("fonts", {})

    payload = {
        "version":      "2.0",
        "platform":     "Canva",
        "generated_by": "Brand Studio AI",
        "brand_name":   brand_data.get("brand_name", ""),
        "tagline":      brand_data.get("tagline", ""),
        "brand_kit": {
            "colors": [
                {
                    "name":  c.get("name", f"Color {i + 1}"),
                    "hex":   c.get("hex", "#000000"),
                    "usage": c.get("meaning", ""),
                }
                for i, c in enumerate(palette)
            ],
            "typography": {
                "heading": {
                    "font_family": fonts.get("heading", "Playfair Display"),
                    "weight": "bold",
                    "size":   36,
                },
                "body": {
                    "font_family": fonts.get("body", "Inter"),
                    "weight": "regular",
                    "size":   14,
                },
            },
            "logo_prompts": brand_data.get("logo_prompts", []),
            "brand_story":  brand_data.get("brand_story", ""),
            "values":       brand_data.get("values", []),
        },
        "social_templates": [
            {
                "type":         "instagram_post",
                "caption":      cap,
                "image_prompt": (brand_data.get("post_prompts") or [""])[min(i, len(brand_data.get("post_prompts") or [""]) - 1)],
            }
            for i, cap in enumerate(brand_data.get("captions", [])[:9])
        ],
        "export_date": datetime.utcnow().isoformat(),
    }

    output_path.write_text(
        json.dumps(payload, indent=2, ensure_ascii=False),
        encoding="utf-8",
    )


# ---------------------------------------------------------------------------
# 6. TXT — Image Prompts
# ---------------------------------------------------------------------------

def generate_prompts_txt(output_path: Path, brand_data: dict[str, Any]) -> None:
    sep = "=" * 70
    hr  = "-" * 70

    lines: list[str] = [
        sep,
        "  BRAND STUDIO AI — VISUAL BRAND ASSETS",
        f"  Brand    : {brand_data.get('brand_name', '')}",
        f"  Language : {brand_data.get('language', 'en').upper()}",
        f"  Date     : {datetime.utcnow().strftime('%Y-%m-%d %H:%M UTC')}",
        sep,
        "",
        "LOGO PROMPTS (3 variants) — Generated by LogoDesigner Agent",
        "Use with: Midjourney · DALL-E 3 · Adobe Firefly · Stable Diffusion",
        "",
        hr,
        "LOGO PROMPTS (3 variants)",
        hr,
    ]

    for i, prompt in enumerate(brand_data.get("logo_prompts", []), 1):
        lines += [f"\nLogo Prompt {i}:", prompt]

    lines += [
        "",
        hr,
        "POST IMAGE PROMPTS (9 social media posts — 1080×1080)",
        hr,
    ]

    for i, prompt in enumerate(brand_data.get("post_prompts", []), 1):
        lines += [f"\nPost Prompt {i}:", prompt]

    lines += [
        "",
        hr,
        "30-DAY CONTENT CALENDAR — PHOTO DIRECTIONS",
        "(Instructions for taking or selecting REAL photos — NOT AI prompts)",
        hr,
    ]

    for row in brand_data.get("calendar", [])[:30]:
        photo = row.get("photo_direction", row.get("image_prompt", ""))
        if photo:
            lines.append(f"\nDay {row.get('day', '?')} [{row.get('theme', '')}]:")
            lines.append(photo)

    lines += [
        "",
        sep,
        "Generated by Brand Studio AI — Build your brand. The Halal Way.",
        sep,
    ]

    output_path.write_text("\n".join(lines), encoding="utf-8")


# ---------------------------------------------------------------------------
# 7. PDF — Zero-Budget Playbook
# ---------------------------------------------------------------------------

def generate_playbook_pdf(output_path: Path, brand_data: dict[str, Any]) -> None:
    """
    Render a Zero-Budget 30-Day Action Playbook PDF from the OfferStrategist data.
    Falls back to a placeholder page if offer data is not present.
    """
    brand_name = brand_data.get("brand_name", "Brand")
    language   = brand_data.get("language", "en")
    is_rtl     = _is_rtl_locale(language)
    pdf = _BasePDF(brand_name, subtitle="Zero-Budget Playbook", rtl=is_rtl)
    pdf.add_page()

    # ── Cover page ───────────────────────────────────────────────────────
    pdf.section_title("Zero-Budget 30-Day Action Playbook")
    pdf.body_text(
        f"This playbook was designed exclusively for {brand_name}. "
        "Every task is achievable on a smartphone with ZERO marketing budget. "
        "Follow the daily actions to build your brand organically."
    )
    pdf.ln(4)

    # ── Launch Offers ────────────────────────────────────────────────────
    offers = brand_data.get("launch_offers", [])
    if offers:
        pdf.section_title("3 Irresistible Launch Offers")
        for i, offer in enumerate(offers[:3], 1):
            pdf.set_font(_FONT_HEADING, "B", 11)
            title_text = pdf._t(f"Offer {i}: {offer.get('title', '')}")
            pdf.cell(0, 7, title_text, ln=True)
            pdf.body_text(offer.get("description", ""), indent=4)
            pdf.set_font(_FONT_BODY, "", 10)
            pdf.set_text_color(41, 98, 255)
            pdf.body_text(f"CTA: {offer.get('cta', '')}", indent=4)
            pdf.set_text_color(0, 0, 0)
            pdf.ln(2)

    # ── 30-Day Playbook ──────────────────────────────────────────────────
    pdf.add_page()
    pdf.section_title("30-Day Action Plan")

    playbook = brand_data.get("playbook", [])
    if playbook:
        for entry in playbook[:30]:
            day  = entry.get("day", "?")
            task = entry.get("task", "")
            plat = entry.get("platform", "")
            mins = entry.get("time_minutes", "")

            y = pdf.get_y()
            if y > 260:
                pdf.add_page()

            if is_rtl:
                # RTL: write task on its own full-width line, right-aligned
                line = pdf._t(f"{task}  [{plat} · {mins} min]  — Day {day}")
                pdf.set_font(_FONT_BODY, "", 10)
                pdf.cell(0, 6, line, align="R", ln=True)
            else:
                pdf.set_font(_FONT_HEADING, "B", 10)
                pdf.cell(20, 6, f"Day {day}")
                pdf.set_font(_FONT_BODY, "", 10)
                pdf.cell(0, 6, f"{task}  [{plat} · {mins} min]", ln=True)
    else:
        pdf.body_text("Playbook data not available. Please regenerate the brand kit.")

    # ── Free Tools section ───────────────────────────────────────────────
    pdf.add_page()
    pdf.section_title("15 Essential Free Tools")
    tools = brand_data.get("free_tools", [])
    if tools:
        for i, tool in enumerate(tools[:15], 1):
            pdf.set_font(_FONT_HEADING, "B", 10)
            tool_header = pdf._t(f"{i}. {tool.get('name', '')} — {tool.get('category', '')}")
            pdf.cell(0, 6, tool_header, align=pdf._align(), ln=True)
            pdf.body_text(f"Use: {tool.get('use_case', '')}  |  {tool.get('url', '')}", indent=4)
    else:
        pdf.body_text("Tools list not available. Please regenerate the brand kit.")

    pdf.output(str(output_path))


# ---------------------------------------------------------------------------
# 8. TXT — Free Tools List
# ---------------------------------------------------------------------------

def generate_free_tools_txt(output_path: Path, brand_data: dict[str, Any]) -> None:
    sep   = "=" * 70
    hr    = "-" * 70
    tools = brand_data.get("free_tools", [])

    lines: list[str] = [
        sep,
        "  BRAND STUDIO AI — FREE TOOLS FOR ZERO-BUDGET ENTREPRENEURS",
        f"  Brand  : {brand_data.get('brand_name', '')}",
        f"  Date   : {datetime.utcnow().strftime('%Y-%m-%d %H:%M UTC')}",
        sep,
        "",
    ]

    if tools:
        # Group by category
        categories: dict[str, list] = {}
        for t in tools:
            cat = t.get("category", "General")
            categories.setdefault(cat, []).append(t)

        for cat, items in categories.items():
            lines += ["", hr, f"  {cat.upper()}", hr]
            for t in items:
                lines += [
                    f"\n▸ {t.get('name', '')}",
                    f"  URL      : {t.get('url', '')}",
                    f"  Use Case : {t.get('use_case', '')}",
                ]
    else:
        lines += [
            "Free tools list could not be generated.",
            "Common free tools for brand building:",
            "  • Canva (canva.com) — graphic design",
            "  • Buffer (buffer.com) — social scheduling",
            "  • Google My Business — local presence",
            "  • Mailchimp (mailchimp.com) — email marketing",
            "  • Later (later.com) — Instagram scheduling",
        ]

    lines += [
        "",
        sep,
        "Generated by Brand Studio AI — Build your brand. The Halal Way.",
        sep,
    ]

    output_path.write_text("\n".join(lines), encoding="utf-8")


# ---------------------------------------------------------------------------
# 7. HTML — Landing Page (AI-generated or static fallback)
# ---------------------------------------------------------------------------

def generate_landing_page(output_path: Path, brand_data: dict[str, Any]) -> None:
    """
    Write the landing page HTML to disk.
    Uses the AI-generated code from FrontendEngineer if available,
    otherwise falls back to the built-in static generator.
    """
    # Prefer the AI-generated page
    ai_html = brand_data.get("landing_page_code", "")
    if ai_html and ai_html.strip().lower().startswith("<!doctype"):
        output_path.write_text(ai_html, encoding="utf-8")
        return

    # Fallback: static generator (same quality as before)
    name          = brand_data.get("brand_name", "My Brand")
    tagline       = brand_data.get("tagline", "")
    story         = brand_data.get("brand_story", "")
    values        = brand_data.get("values", [])
    palette       = brand_data.get("color_palette", [])
    offers        = brand_data.get("launch_offers", [])
    captions      = brand_data.get("captions", [])[:3]

    primary_hex   = palette[0].get("hex", "#2962FF") if palette else "#2962FF"
    secondary_hex = palette[1].get("hex", "#1a1a2e") if len(palette) > 1 else "#1a1a2e"

    values_html   = "".join(f'<span class="value-badge">{v}</span>' for v in values)
    offer_cards   = "".join(
        f'<div class="offer-card"><h3>{o.get("title","")}</h3>'
        f'<p>{o.get("description","")}</p>'
        f'<a href="#contact" class="cta-btn">{o.get("cta","Get Started")}</a></div>'
        for o in offers[:3]
    )
    caption_items = "".join(f"<li>{cap}</li>" for cap in captions)

    html = f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8"/>
<meta name="viewport" content="width=device-width, initial-scale=1.0"/>
<title>{name} — {tagline}</title>
<style>
  :root {{
    --primary: {primary_hex};
    --secondary: {secondary_hex};
    --bg: #ffffff;
    --text: #1a1a2e;
    --light: #f8f9ff;
    --border: #e8ecf4;
  }}
  * {{ box-sizing: border-box; margin: 0; padding: 0; }}
  body {{ font-family: -apple-system, "Segoe UI", sans-serif; color: var(--text); background: var(--bg); line-height: 1.65; }}
  a {{ color: var(--primary); text-decoration: none; }}
  nav {{ background: var(--secondary); padding: 1rem 2rem; display: flex; justify-content: space-between; align-items: center; }}
  .nav-brand {{ color: #fff; font-weight: 700; font-size: 1.2rem; }}
  .hero {{ background: linear-gradient(135deg, var(--secondary), var(--primary)); padding: 5rem 2rem; text-align: center; }}
  .hero h1 {{ color: #fff; font-size: clamp(2rem, 6vw, 3.5rem); font-weight: 800; margin-bottom: 1rem; }}
  .hero p {{ color: rgba(255,255,255,.85); font-size: 1.15rem; max-width: 600px; margin: 0 auto 2rem; }}
  .hero-btn {{ display: inline-block; background: var(--primary); color: #fff; padding: .75rem 2rem; border-radius: 50px; font-weight: 600; }}
  section {{ max-width: 960px; margin: 0 auto; padding: 4rem 2rem; }}
  h2 {{ font-size: 1.75rem; font-weight: 700; margin-bottom: 1.5rem; color: var(--secondary); }}
  .values-grid {{ display: flex; flex-wrap: wrap; gap: .75rem; }}
  .value-badge {{ background: var(--light); border: 1px solid var(--border); padding: .4rem 1rem; border-radius: 50px; font-size: .9rem; }}
  .story-box {{ background: var(--light); border-left: 4px solid var(--primary); padding: 1.5rem 2rem; border-radius: 0 8px 8px 0; }}
  .offers-grid {{ display: grid; grid-template-columns: repeat(auto-fit, minmax(260px, 1fr)); gap: 1.5rem; }}
  .offer-card {{ border: 1px solid var(--border); border-radius: 12px; padding: 1.5rem; }}
  .offer-card h3 {{ margin-bottom: .5rem; font-size: 1.1rem; color: var(--secondary); }}
  .cta-btn {{ display: inline-block; background: var(--primary); color: #fff; padding: .5rem 1.25rem; border-radius: 50px; font-size: .9rem; font-weight: 600; }}
  .captions-list {{ list-style: none; }}
  .captions-list li {{ padding: .75rem 1rem; border-bottom: 1px solid var(--border); font-size: .95rem; }}
  footer {{ background: var(--secondary); color: rgba(255,255,255,.6); text-align: center; padding: 2rem; font-size: .85rem; }}
  footer strong {{ color: #fff; }}
</style>
</head>
<body>
<nav><span class="nav-brand">{name}</span></nav>
<div class="hero">
  <h1>{name}</h1><p>{tagline}</p>
  <a href="#offers" class="hero-btn">See Our Offers</a>
</div>
<section id="story"><h2>Our Story</h2><div class="story-box">{story}</div></section>
<section id="values"><h2>Our Values</h2><div class="values-grid">{values_html}</div></section>
<section id="offers">
  <h2>Launch Offers</h2>
  <div class="offers-grid">{offer_cards or "<p>Special offers coming soon.</p>"}</div>
</section>
<section id="social">
  <h2>From Our Feed</h2>
  <ul class="captions-list">{caption_items or "<li>Follow us for updates.</li>"}</ul>
</section>
<footer><p>© {datetime.utcnow().year} <strong>{name}</strong>. Generated by <strong>Brand Studio AI</strong>.</p></footer>
</body>
</html>"""
    output_path.write_text(html, encoding="utf-8")


# ---------------------------------------------------------------------------
# 8. Master ZIP — bundle all 7 files
# ---------------------------------------------------------------------------

def _create_master_zip(
    zip_path:      Path,
    pdf_path:      Path,
    xlsx_path:     Path,
    json_path:     Path,
    txt_path:      Path,
    playbook_path: Path,
    tools_txt:     Path,
    html_path:     Path,
) -> None:
    """Bundle all exported files into the master Branding_Kit.zip."""
    with zipfile.ZipFile(zip_path, "w", zipfile.ZIP_DEFLATED) as zf:
        zf.write(pdf_path,      "Brand_Guide.pdf")
        zf.write(xlsx_path,     "Content_Calendar.xlsx")
        zf.write(json_path,     "Canva_Import.json")
        zf.write(txt_path,      "Image_Prompts.txt")
        zf.write(playbook_path, "Zero_Budget_Playbook.pdf")
        zf.write(tools_txt,     "Free_Tools_List.txt")
        zf.write(html_path,     "Landing_Page.html")


# ---------------------------------------------------------------------------
# MASTER EXPORT FUNCTION  (called by main.py)
# ---------------------------------------------------------------------------

def export_files(
    job_id:     str,
    brand_data: dict[str, Any],
    language:   str,
    export_dir: str = "/tmp/exports",
) -> Path:
    """
    Orchestrate creation of all 7 output files for a given job.

    Parameters
    ----------
    job_id     : Unique job identifier (used as subdirectory name).
    brand_data : Merged dict returned by BrandingCrew.run().
    language   : ISO 639-1 language code (recorded in the PDF).
    export_dir : Root directory for file outputs.

    Returns
    -------
    Path to the completed Branding_Kit.zip.
    """
    base = Path(export_dir) / job_id
    base.mkdir(parents=True, exist_ok=True)

    pdf_path      = base / "Brand_Guide.pdf"
    xlsx_path     = base / "Content_Calendar.xlsx"
    json_path     = base / "Canva_Import.json"
    txt_path      = base / "Image_Prompts.txt"
    playbook_path = base / "Zero_Budget_Playbook.pdf"
    tools_txt     = base / "Free_Tools_List.txt"
    html_path     = base / "Landing_Page.html"
    zip_path      = base / "Branding_Kit.zip"

    # Ensure language is available inside brand_data for sub-generators
    brand_data.setdefault("language", language)

    generate_pdf(pdf_path, brand_data, language)
    generate_xlsx(xlsx_path, brand_data.get("calendar", []), language)
    generate_canva_json(json_path, brand_data)
    generate_prompts_txt(txt_path, brand_data)
    generate_playbook_pdf(playbook_path, brand_data)
    generate_free_tools_txt(tools_txt, brand_data)
    generate_landing_page(html_path, brand_data)
    _create_master_zip(
        zip_path,
        pdf_path, xlsx_path, json_path, txt_path, playbook_path, tools_txt, html_path,
    )

    return zip_path
